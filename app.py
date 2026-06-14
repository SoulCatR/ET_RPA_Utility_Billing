import streamlit as st
import pdfplumber
import openpyxl
import re
import io
import zipfile
import pandas as pd

# ==========================================
# CONFIGURACIÓN Y MAPEOS
# ==========================================
COL_PERIODO = 1      # Columna A
COL_LOCAL = 3        # Columna C

MAPEO_COLUMNAS = {
    "840": {"cant": 6, "prec": 7},   # Agua
    "841": {"cant": 9, "prec": 10},  # Energía
    "824": {"cant": 12, "prec": 13}, # Aire Acond.
    "842": {"cant": 15, "prec": 16}  # Emergencia
}

TIENDAS_KEYWORDS = {
    "Almacen Retail - Sotano Airside": ["almacén retail - sotano airside", "almacen retail - sotano airside", "sotano airside"],
    "Almacen Retail": ["almacén retail", "almacen retail"],
    "Grab&Go Domestico": ["grab&go domestico", "grab & go domestico"],
    "Grab&Go Dique": ["grab&go dique", "grab & go dique"],
    "Kaphy & Kakau": ["kaphy & kakau", "kaphy", "kakau"],
    "Oasis": ["oasis del sol", "oasis"],
    "Pharmacy": ["farmacia morpho", "pharmacy", "farmacia"],
    "Travel": ["travel", "salidas nacionales"], 
    "Retablo": ["retablo"],
    "Expansión": ["convenience check in 2 britt", "convenience check in 2", "expansión", "expansion"],
    "Chakana": ["chakana"],
    "Quipu": ["quipu"],
    "Oficina": ["oficina"]
}

# ==========================================
# FUNCIONES AUXILIARES
# ==========================================
def limpiar_numero(texto):
    if not texto: return 0.0
    try:
        return float(texto.replace(",", "").strip())
    except ValueError:
        return 0.0

def identificar_tienda(texto):
    texto_lower = texto.lower()
    for tienda_excel, keywords in TIENDAS_KEYWORDS.items():
        if any(kw in texto_lower for kw in keywords):
            return tienda_excel
    return None

# ==========================================
# INTERFAZ WEB CON STREAMLIT
# ==========================================
st.set_page_config(page_title="Consumos a Excel", layout="wide")
st.title("⚡ Procesador de Consumos - Agua y Energía")
st.markdown("Sube tu archivo maestro de Excel y los recibos en PDF. Puedes subir múltiples PDFs a la vez o un archivo `.zip` que los contenga.")

col1, col2 = st.columns(2)
with col1:
    archivo_excel = st.file_uploader("1. Sube la plantilla de Excel (.xlsx)", type=["xlsx"])
with col2:
    archivos_pdf_subidos = st.file_uploader("2. Sube los PDFs o un archivo .ZIP", type=["pdf", "zip"], accept_multiple_files=True)

if st.button("🚀 Procesar Datos", type="primary"):
    if not archivo_excel or not archivos_pdf_subidos:
        st.warning("⚠️ Por favor, sube el Excel y al menos un PDF/ZIP para comenzar.")
    else:
        with st.spinner("Analizando PDFs e inyectando datos..."):
            
            # 1. Cargar el Excel en memoria
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            mapa_filas = {}
            mes_actual = None

            for row in range(8, ws.max_row + 1):
                valor_periodo = ws.cell(row=row, column=COL_PERIODO).value
                if valor_periodo:
                    mes_actual = str(valor_periodo).strip().lower()
                valor_local = ws.cell(row=row, column=COL_LOCAL).value
                if valor_local and mes_actual:
                    local_limpio = str(valor_local).strip()
                    mapa_filas[(mes_actual, local_limpio)] = row

            # 2. Desempaquetar archivos (Manejo de PDFs y ZIPs)
            archivos_a_procesar = [] # Lista de tuplas (nombre_archivo, bytes_del_pdf)
            for file in archivos_pdf_subidos:
                if file.name.lower().endswith(".zip"):
                    with zipfile.ZipFile(file) as z:
                        for filename in z.namelist():
                            if filename.lower().endswith(".pdf"):
                                # Se evita leer carpetas de macOS como __MACOSX
                                if "__MACOSX" not in filename:
                                    archivos_a_procesar.append((os.path.basename(filename), io.BytesIO(z.read(filename))))
                else:
                    archivos_a_procesar.append((file.name, file))

            # 3. Procesamiento
            log_resultados = []
            
            for filename, file_bytes in archivos_a_procesar:
                match = re.search(r'([AaEe])\s*-\s*([^-]+)\s*-\s*(.*?)\.pdf$', filename)
                if not match:
                    log_resultados.append({"ARCHIVO": filename, "ESTADO": "OMITIDO", "TIENDAS": "-", "DETALLE": "Formato de nombre no reconocido"})
                    continue

                tipo = match.group(1).upper()
                mes = match.group(2).lower().strip()

                try:
                    with pdfplumber.open(file_bytes) as pdf:
                        tablas_extraidas = False
                        datos_consolidados = {}
                        tiendas_en_este_pdf = set()

                        for page in pdf.pages:
                            for tabla in page.extract_tables():
                                if not tabla: continue
                                tablas_extraidas = True
                                header_row_idx = -1
                                idx_codigo = idx_desc = idx_cant = idx_prec = idx_item = None
                                
                                for row_idx, row in enumerate(tabla[:5]):
                                    headers = [str(h).replace('\n', ' ').strip().lower() if h else '' for h in row]
                                    i_cant = next((i for i, h in enumerate(headers) if 'cantidad' in h or 'cant' in h), None)
                                    i_prec = next((i for i, h in enumerate(headers) if 'precio' in h or 'unitario' in h), None)
                                    
                                    if i_cant is not None and i_prec is not None:
                                        idx_cant = i_cant
                                        idx_prec = i_prec
                                        idx_codigo = next((i for i, h in enumerate(headers) if 'código' in h or 'codigo' in h or 'cod' in h or 'cód' in h), None)
                                        idx_desc = next((i for i, h in enumerate(headers) if 'descripción' in h or 'descripcion' in h or 'detalle' in h or 'concepto' in h), None)
                                        idx_item = next((i for i, h in enumerate(headers) if 'item' in h), None)
                                        header_row_idx = row_idx
                                        break
                                
                                if header_row_idx == -1: continue

                                for num_fila_interna, fila in enumerate(tabla[header_row_idx+1:], start=1):
                                    if not fila: continue
                                    
                                    if idx_codigo is not None and len(fila) > idx_codigo and fila[idx_codigo]:
                                        codigo = str(fila[idx_codigo]).strip()
                                    else:
                                        if tipo == 'A': codigo = "840" 
                                        else: continue 

                                    if codigo not in MAPEO_COLUMNAS: continue

                                    desc = str(fila[idx_desc]).replace('\n', ' ').strip() if idx_desc is not None and len(fila) > idx_desc and fila[idx_desc] else ""
                                    num_item = str(fila[idx_item]).strip() if idx_item is not None and len(fila) > idx_item and fila[idx_item] else str(num_fila_interna)
                                    cant = limpiar_numero(fila[idx_cant]) if len(fila) > idx_cant else 0.0
                                    prec = limpiar_numero(fila[idx_prec]) if len(fila) > idx_prec else 0.0

                                    tienda_excel = identificar_tienda(desc)

                                    if not tienda_excel:
                                        log_resultados.append({"ARCHIVO": filename, "ESTADO": "NO UBICADA", "TIENDAS": ", ".join(tiendas_en_este_pdf), "DETALLE": f"Item {num_item} sin tienda identificada"})
                                        continue

                                    tiendas_en_este_pdf.add(tienda_excel)

                                    if tienda_excel not in datos_consolidados:
                                        datos_consolidados[tienda_excel] = {}
                                    if codigo not in datos_consolidados[tienda_excel]:
                                        datos_consolidados[tienda_excel][codigo] = {'cant': 0.0, 'prec': prec}

                                    datos_consolidados[tienda_excel][codigo]['cant'] += cant
                                    datos_consolidados[tienda_excel][codigo]['prec'] = prec 

                        if not tablas_extraidas:
                            log_resultados.append({"ARCHIVO": filename, "ESTADO": "ERROR", "TIENDAS": "-", "DETALLE": "No se encontraron tablas"})
                            continue
                        if not datos_consolidados:
                            log_resultados.append({"ARCHIVO": filename, "ESTADO": "SIN DATOS", "TIENDAS": "-", "DETALLE": "Sin códigos válidos (840, 841, etc.)"})
                            continue 

                        exito_escritura = False
                        for tienda_excel, codigos in datos_consolidados.items():
                            fila_objetivo = None
                            for (m, t), fila_idx in mapa_filas.items():
                                if m == mes and t.lower() == tienda_excel.lower():
                                    fila_objetivo = fila_idx
                                    break

                            if not fila_objetivo:
                                log_resultados.append({"ARCHIVO": filename, "ESTADO": "FILA NO HALLADA", "TIENDAS": ", ".join(tiendas_en_este_pdf), "DETALLE": f"Tienda '{tienda_excel}' (Mes: {mes}) no hallada"})
                                continue

                            for codigo, valores in codigos.items():
                                if codigo in MAPEO_COLUMNAS:
                                    ws.cell(row=fila_objetivo, column=MAPEO_COLUMNAS[codigo]['cant'], value=valores['cant'])
                                    ws.cell(row=fila_objetivo, column=MAPEO_COLUMNAS[codigo]['prec'], value=valores['prec'])
                                    exito_escritura = True

                        if exito_escritura:
                            log_resultados.append({"ARCHIVO": filename, "ESTADO": "ÉXITO", "TIENDAS": ", ".join(tiendas_en_este_pdf), "DETALLE": "Data insertada correctamente"})

                except Exception as e:
                    log_resultados.append({"ARCHIVO": filename, "ESTADO": "ERROR LECTURA", "TIENDAS": "-", "DETALLE": f"Fallo: {str(e)[:30]}"})

            # 4. Guardar resultado final en memoria y mostrar
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            st.success(f"¡Se procesaron {len(archivos_a_procesar)} archivos PDF correctamente!")
            
            # Botón verde gigante de descarga
            st.download_button(
                label="📥 Descargar Excel Actualizado",
                data=output,
                file_name="Consumo_Recursos_Actualizado.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )

            # Mostrar tabla en la web
            st.write("### 📋 Log de Ejecución")
            df_log = pd.DataFrame(log_resultados)
            st.dataframe(df_log, use_container_width=True)